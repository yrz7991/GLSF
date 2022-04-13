from playsound import playsound
import librosa
import os
import openpyxl
path = r"C:\Users\zry\PycharmProjects\recommendation\music\14"
file_path = os.listdir(path)
# ws = openpyxl.Workbook()
# sheet = ws.active
# ws.create_sheet(index=0, title='tempo')
# ws.create_sheet(index=1, title='rms')
# ws.create_sheet(index=2, title='zero')
# # 保存工作表
# ws.save("feature.xlsx")
num = 0
print(file_path)
for filename in file_path:
    num = num + 1
    path_filename = os.path.join("./music/14", filename)  # ./代表本目录
    y, sr = librosa.load(path_filename)
    tempo, beats = librosa.beat.beat_track(y=y, sr=sr)  # 节奏和与节拍对应的帧
    time = librosa.frames_to_time(beats, sr=sr)  # 将与节拍对应的帧转换为秒
    tempo_be = []  # 用于存储每两节拍之间的速度
    # wb = openpyxl.load_workbook('feature.xlsx')
    # sheet_tempo = wb.worksheets[0]
    # sheet_tempo.cell(num, 1).value = filename
    # sheet_rms = wb.worksheets[1]
    # sheet_rms.cell(num, 1).value = filename
    # sheet_zero = wb.worksheets[2]
    # sheet_zero.cell(num, 1).value = filename
    i = 0
    num_m = 1
    while i < len(time) - 1:
        num_m = num_m + 1
        con_time = time[i+1] - time[i]
        if i == 0:
            set_time = 0
            pass
        else:
            set_time = time[i-1]
        y_be, sr_be = librosa.load(path_filename, offset=set_time, duration=con_time)
        tempo_be_c, beats_be_c = librosa.beat.beat_track(y=y_be, sr=sr_be)
        # sheet_tempo.cell(num_m, num - 1).value = 'tempo'
        # sheet_tempo.cell(num, num_m).value = tempo_be_c
        # wb.save('feature.xlsx')
        tempo_be.append(tempo_be_c)
        i = i+1
        pass
    # print(tempo_be)
    rms_be = []  # 用于存储每两节拍之间的rms值
    zero_be = []  # 用于存储每两节拍之间的过零率
    cent_be = []
    spec_bw_be = []
    flatness_be = []
    S, phase = librosa.magphase(librosa.stft(y))
    rms = librosa.feature.rms(S=S)  # 从频谱图输入获取rms的值(会提供更准确的能量表示)，输出为每帧的rms值
    zero = librosa.feature.zero_crossing_rate(y)  # 过零率
    cent = librosa.feature.spectral_centroid(y=y, sr=sr)  # 光谱质心
    spec_bw = librosa.feature.spectral_bandwidth(y=y, sr=sr)  # 光谱带宽
    flatness = librosa.feature.spectral_flatness(y=y)  # 光谱平坦度
    j = 0
    num_m = 1
    while j < len(beats) - 1:
        k = 0
        sum_rms = 0
        sum_zero = 0
        sum_cent = 0
        sum_spec_bw = 0
        sum_flatness = 0
        num_m = num_m + 1
        if j == 0:
            while k < beats[0]:
                sum_rms += rms[0][k]
                sum_zero += zero[0][k]
                sum_cent += cent[0][k]
                sum_spec_bw += spec_bw[0][k]
                sum_flatness += flatness[0][k]
                k = k + 1
                pass
            mean_rms = sum_rms / beats[0]
            mean_zero = sum_zero / beats[0]
            mean_cent = sum_cent / beats[0]
            mean_spec_bw = sum_spec_bw / beats[0]
            mean_flatness = sum_flatness / beats[0]
            pass
        else:
            while k < beats[j] - beats[j-1]:
                sum_rms += rms[0][beats[j-1]+1]
                sum_zero += zero[0][beats[j-1]+1]
                sum_cent += cent[0][beats[j-1]+1]
                sum_spec_bw += spec_bw[0][beats[j-1]+1]
                sum_flatness += flatness[0][beats[j-1]+1]
                k = k + 1
                pass
            mean_rms = sum_rms / k
            mean_zero = sum_zero / k
            mean_cent = sum_cent / k
            mean_spec_bw = sum_spec_bw / k
            mean_flatness = sum_flatness / k
            pass
        j = j + 1
        rms_be.append(mean_rms)
        zero_be.append(mean_zero)
        cent_be.append(mean_cent)
        spec_bw_be.append(mean_spec_bw)
        flatness_be.append(mean_flatness)
        # sheet_rms.cell(num, num_m).value = mean_rms
        # sheet_zero.cell(num, num_m).value = mean_zero
        # wb.save('feature.xlsx')
        pass
    print('tempo')
    print(tempo_be)
    print('rms')
    print(rms_be)
    print('zero')
    print(zero_be)
    print('cent')
    print(cent_be)
    print('spec_bw')
    print(spec_bw_be)
    print('flatness')
    print(flatness_be)
    # print(len(rms_be))
    # print(len(tempo_be))
    # print(len(zero_be))

