#from playsound import playsound
import librosa
import os
import openpyxl
import madmom
from madmom.features.beats import RNNBeatProcessor
from madmom.features.tempo import TempoEstimationProcessor
path = r"C:\Users\hello\Desktop\recommendation\music\10"
file_path = os.listdir(path)
ws = openpyxl.Workbook()
sheet = ws.active
ws.create_sheet(index=0, title='tempo')
ws.create_sheet(index=1, title='rms')
ws.create_sheet(index=2, title='zero')
# 保存工作表
ws.save("feature1.xlsx")
num = 0
print(file_path)
def sum_avg(list):
    sum = 0
    for n in list:
        sum += n
    return sum / len(list)

def ex(list):
    sum = 0
    for i in range(0,len(list)):
        sum += list[i][0]*list[i][1]
    return sum

for filename in file_path:
    num = num + 1
    path_filename = os.path.join("./music/10", filename)  # ./代表本目录
    y, sr = librosa.load(path_filename)
    tempo, beats = librosa.beat.beat_track(y=y, sr=sr)  # 节奏和与节拍对应的帧
    wb = openpyxl.load_workbook('feature1.xlsx')
    sheet_tempo = wb.worksheets[0]
    sheet_tempo.cell(num, 1).value = filename
    sheet_rms = wb.worksheets[1]
    sheet_rms.cell(num, 1).value = filename
    sheet_zero = wb.worksheets[2]
    sheet_zero.cell(num, 1).value = filename
    proc = TempoEstimationProcessor(fps=100)
    act = RNNBeatProcessor()(path_filename)  
    sheet_tempo.cell(num, 2).value = ex(proc(act))
    print(ex(proc(act)))
    # wb.save('feature.xlsx')
    S, phase = librosa.magphase(librosa.stft(y))
    rms = librosa.feature.rms(S=S)  # 从频谱图输入获取rms的值(会提供更准确的能量表示)，输出为每帧的rms值
    rms1 = sum_avg(rms[0])
    zero = librosa.feature.zero_crossing_rate(y)  # 过零率
    zero1 = sum_avg(zero[0])
    sheet_rms.cell(num, 2).value = rms1
    sheet_zero.cell(num, 2).value = zero1
    wb.save('feature1.xlsx')

