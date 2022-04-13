# -*- coding: utf-8 -*-
"""
Created on Tue Dec 21 19:54:59 2021

@author: hello
"""

import librosa
import os
import openpyxl
import datetime
import madmom
from madmom.features.beats import RNNBeatProcessor
from madmom.features.tempo import TempoEstimationProcessor
path = r"C:\Users\hello\Desktop\recommendation\music\11"
file_path = os.listdir(path)
ws = openpyxl.Workbook()
sheet = ws.active
ws.create_sheet(index=0, title='tempo')
ws.create_sheet(index=1, title='rms')
ws.create_sheet(index=2, title='zero')
# 保存工作表
ws.save("feature.xlsx")
num = 0
print(file_path)

def ex(list):
    sum = 0
    for i in range(0,len(list)):
        sum += list[i][0]*list[i][1]
    return sum
t1 = datetime.datetime.now()
for filename in file_path:
    num = num + 1
    path_filename = os.path.join("./music/11", filename)  # ./代表本目录
    y, sr = librosa.load(path_filename)
    tempo, beats = librosa.beat.beat_track(y=y, sr=sr)  # 节奏和与节拍对应的帧
    # time = librosa.frames_to_time(beats, sr=sr)  # 将与节拍对应的帧转换为秒
    # tempo_be = []  # 用于存储每两节拍之间的速度
    # wb = openpyxl.load_workbook('feature.xlsx')
    # sheet_tempo = wb.worksheets[0]
    # sheet_tempo.cell(num, 1).value = filename
    # sheet_rms = wb.worksheets[1]
    # sheet_rms.cell(num, 1).value = filename
    # sheet_zero = wb.worksheets[2]
    # sheet_zero.cell(num, 1).value = filename
    # i = 0
    # num_m = 1
    # t1 = datetime.datetime.now()
    proc = TempoEstimationProcessor(fps=100)
    act = RNNBeatProcessor()(path_filename)  
    # sheet_tempo.cell(num, 2).value = ex(proc(act))
    print(ex(proc(act)))
    # wb.save('feature.xlsx')
    # while i < len(time) - 1:
    #     num_m = num_m + 1
    #     con_time = time[i+1] - time[i]
    #     if i == 0:
    #         set_time = 0
    #         pass
    #     else:
    #         set_time = time[i-1]
    #     y_be, sr_be = librosa.load(path_filename, offset=set_time, duration=con_time)
    #     tempo_be_c, beats_be_c = librosa.beat.beat_track(y=y_be, sr=sr_be)
        # sheet_tempo.cell(num_m, num - 1).value = 'tempo'
        # sheet_tempo.cell(num, num_m).value = tempo_be_c
        # wb.save('feature.xlsx')
        # tempo_be.append(tempo_be_c)
        # i = i+1
        # pass
    # print(tempo_be)
    rms_be = []  # 用于存储每两节拍之间的rms值
    zero_be = []  # 用于存储每两节拍之间的过零率
    S, phase = librosa.magphase(librosa.stft(y))
    rms = librosa.feature.rms(S=S)  # 从频谱图输入获取rms的值(会提供更准确的能量表示)，输出为每帧的rms值
    zero = librosa.feature.zero_crossing_rate(y)  # 过零率
    j = 0
    num_m = 1
    while j < len(beats) - 1:
        k = 0
        sum_rms = 0
        sum_zero = 0
        num_m = num_m + 1
        if j == 0:
            while k < beats[0]:
                sum_rms += rms[0][k]
                sum_zero += zero[0][k]
                k = k + 1
                pass
            mean_rms = sum_rms / beats[0]
            mean_zero = sum_zero / beats[0]
            pass
        else:
            while k < beats[j] - beats[j-1]:
                sum_rms += rms[0][beats[j-1]+1]
                sum_zero += zero[0][beats[j-1]+1]
                k = k + 1
                pass
            mean_rms = sum_rms / k
            mean_zero = sum_zero / k
            pass
        j = j + 1
        rms_be.append(mean_rms)
        zero_be.append(mean_zero)
        # sheet_rms.cell(num_m, num - 1).value = 'mean'
        # sheet_rms.cell(num, num_m).value = mean_rms
        # sheet_zero.cell(num, num_m).value = mean_zero
        # wb.save('feature.xlsx')
        pass
t2 = datetime.datetime.now()
print(t2 - t1)
    # print(rms_be)
    # print(zero_be)
    # print(len(rms_be))
    # # print(len(tempo_be))
    # print(len(zero_be))
